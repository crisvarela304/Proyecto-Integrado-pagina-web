from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count
from django import forms
from django.forms import modelform_factory
import secrets

from academico.models import Curso, Asignatura
from .services import LiceoOSService

def es_administrativo_check(user):
    """Verifica si el usuario es staff o tiene perfil administrativo/directivo"""
    if user.is_staff:
        return True
    if hasattr(user, 'perfil'):
        return user.perfil.tipo_usuario in ['administrativo', 'directivo']
    return False

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def dashboard_main(request):
    """
    Vista principal del Panel de Gestión Integral (LiceoOS).
    Solo accesible para Administrativos y Directivos.
    """
    try:
        # 1. Obtener KPIs en tiempo real
        kpis = LiceoOSService.get_kpis_globales()
        
        # 2. Ejecutar Sistema de Alerta Temprana (SAT)
        alertas = LiceoOSService.detectar_riesgo_academico()
        
        context = {
            'kpis': kpis,
            'alertas': alertas, # Lista de diccionarios con riesgo
            'page_title': 'LiceoOS - Centro de Operaciones',
            'actividades_recientes': LiceoOSService.get_historial_actividad(limite=8)
        }
        return render(request, 'administrativo/dashboard.html', context)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        # Fallback seguro en caso de error en el servicio
        messages.error(request, f"Error cargando el dashboard inteligente: {str(e)}")
        return render(request, 'administrativo/dashboard_error.html', {})

# --- Módulo: Gestión Académica (Cursos) ---

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def gestion_cursos(request):
    """Lista visual de cursos para administración"""
    # Parametros de filtro
    anio_seleccionado = request.GET.get('anio')
    busqueda = request.GET.get('q', '').strip()
    
    # Queryset base
    cursos = Curso.objects.annotate(
        total_estudiantes_real=Count('estudiantes')
    )
    
    # Aplicar filtros
    if anio_seleccionado and anio_seleccionado.isdigit():
        cursos = cursos.filter(año=anio_seleccionado)
        
    if busqueda:
        cursos = cursos.filter(
            Q(nombre__icontains=busqueda) | 
            Q(letra__iexact=busqueda) |
            Q(profesor_jefe__first_name__icontains=busqueda) |
            Q(profesor_jefe__last_name__icontains=busqueda)
        )
        
    # Ordenamiento
    cursos = cursos.order_by('nivel', 'letra')

    # Obtener años disponibles para el filtro
    anios_qs = Curso.objects.values_list('año', flat=True).distinct().order_by('-año')
    anios_list = []
    target_anio = int(anio_seleccionado) if anio_seleccionado and anio_seleccionado.isdigit() else None
    
    for anio in anios_qs:
        anios_list.append({
            'valor': anio,
            'is_selected': (anio == target_anio)
        })

    return render(request, 'administrativo/gestion_cursos.html', {
        'cursos': cursos,
        'page_title': 'Gestión de Cursos',
        'anios_list': anios_list,
        'busqueda': busqueda
    })

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def curso_crear(request):
    """Formulario de creación de curso"""
    CursoForm = modelform_factory(Curso, fields=['nivel', 'letra', 'profesor_jefe', 'activo', 'año'], 
                                 widgets={
                                     'nivel': forms.Select(attrs={'class': 'form-select'}),
                                     'letra': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'A'}),
                                     'profesor_jefe': forms.Select(attrs={'class': 'form-select'}),
                                     'año': forms.NumberInput(attrs={'class': 'form-control'}),
                                     'activo': forms.CheckboxInput(attrs={'class': 'form-check-input'})
                                 })
    
    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            curso = form.save()
            # Auto-asignar nombre (Ej: 1° Medio A)
            nivel_dict = dict(Curso.NIVEL_CHOICES)
            curso.nombre = f"{nivel_dict.get(curso.nivel, '')} {curso.letra}"
            curso.save()
            messages.success(request, f'Curso "{curso.nombre}" creado exitosamente.')
            return redirect('administrativo:gestion_cursos')
    else:
        form = CursoForm()
        
    return render(request, 'administrativo/form_modal_generico.html', {
        'form': form,
        'titulo': 'Crear Nuevo Curso',
        'boton_texto': 'Crear Curso',
        'icono': 'bi-plus-circle'
    })

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def curso_editar(request, pk):
    """Formulario de edición de curso"""
    curso = get_object_or_404(Curso, pk=pk)
    CursoForm = modelform_factory(Curso, fields=['nivel', 'letra', 'profesor_jefe', 'activo', 'año'], 
                                 widgets={
                                     'nivel': forms.Select(attrs={'class': 'form-select'}),
                                     'letra': forms.TextInput(attrs={'class': 'form-control'}),
                                     'profesor_jefe': forms.Select(attrs={'class': 'form-select'}),
                                     'año': forms.NumberInput(attrs={'class': 'form-control'}),
                                     'activo': forms.CheckboxInput(attrs={'class': 'form-check-input'})
                                 })
    
    if request.method == 'POST':
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            curso = form.save()
            # Actualizar nombre
            nivel_dict = dict(Curso.NIVEL_CHOICES)
            curso.nombre = f"{nivel_dict.get(curso.nivel, '')} {curso.letra}"
            curso.save()
            messages.success(request, 'Curso actualizado correctamente.')
            return redirect('administrativo:gestion_cursos')
    else:
        form = CursoForm(instance=curso)
        
    return render(request, 'administrativo/form_modal_generico.html', {
        'form': form,
        'titulo': f'Editar {curso.nombre}',
        'boton_texto': 'Guardar Cambios',
        'icono': 'bi-pencil'
    })

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def curso_eliminar(request, pk):
    """Eliminación segura de curso"""
    curso = get_object_or_404(Curso, pk=pk)
    if request.method == 'POST':
        nombre = curso.nombre
        curso.delete()
        messages.success(request, f'Curso "{nombre}" eliminado correctamente.')
        return redirect('administrativo:gestion_cursos')
    
    return render(request, 'administrativo/confirmar_eliminar.html', {
        'objeto': curso,
        'tipo': 'Curso',
        'back_url': 'administrativo:gestion_cursos'
    })

# --- Carga Masiva de Alumnos ---


# --- Carga Masiva de Alumnos y Profesores ---

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def descargar_plantilla_carga(request, tipo='estudiante'):
    """Genera un archivo Excel de ejemplo para la carga masiva"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Plantilla {tipo.capitalize()}s"

    # Encabezados
    if tipo == 'profesor':
        headers = ['RUT', 'Nombres', 'Apellidos', 'Email']
        filename = 'plantilla_profesores_liceoos.xlsx'
    else:
        headers = ['RUT', 'Nombres', 'Apellidos', 'Email', 'Curso']
        filename = 'plantilla_alumnos_liceoos.xlsx'
        
    ws.append(headers)

    # Estilo encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Ejemplo de datos
    if tipo == 'profesor':
        ws.append(['11222333-4', 'Carlos', 'Muñoz', 'carlos.munoz@liceo.cl'])
        ws.append(['99888777-K', 'Ana', 'Rojas', 'ana.rojas@liceo.cl'])
    else:
        ws.append(['12345678-9', 'Juan Pablo', 'Pérez González', 'juan.perez@liceo.cl', '1° Medio A'])
        ws.append(['87654321-K', 'Maria Paz', 'Soto Soto', 'maria.soto@liceo.cl', '2° Medio B'])

    # Ajustar ancho columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    if tipo != 'profesor':
        ws.column_dimensions['E'].width = 15

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def carga_masiva_estudiantes(request):
    """Procesa la carga masiva de estudiantes desde Excel"""
    from .forms import CargaMasivaForm
    import openpyxl
    from django.contrib.auth.models import User
    from usuarios.models import PerfilUsuario
    from academico.models import Curso, InscripcionCurso
    from core.utils import validar_rut 

    if request.method == 'POST':
        form = CargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active

            creados = 0
            errores = 0
            errores_log = []

            # Iterar filas (saltando encabezado)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Protección contra filas vacías o incompletas
                if not row or all(cell is None for cell in row):
                    continue
                    
                # Desempaquetado seguro
                try:
                    if len(row) >= 5:
                        rut, nombres, apellidos, email, nombre_curso = row[:5]
                    else:
                        raise ValueError("Fila con menos columnas de las esperadas")
                except ValueError as ve:
                    # Intento de manejo si faltan columnas
                    continue
                
                # Validaciones básicas de fila vacía
                if not rut or not nombres or not apellidos:
                    continue
                
                rut = str(rut).strip().upper().replace('.', '')
                
                try:
                    # 1. Verificar si usuario ya existe
                    if User.objects.filter(username=rut).exists():
                         # Opcional: Actualizar datos
                        pass
                    else:
                        # 2. Crear Usuario (Password segura)
                        raw_password = secrets.token_urlsafe(8)
                        user = User.objects.create_user(
                            username=rut,
                            email=email or '',
                            password=raw_password,
                            first_name=nombres,
                            last_name=apellidos
                        )
                        
                        # 3. Crear Perfil
                        PerfilUsuario.objects.create(
                            user=user,
                            rut=rut,
                            tipo_usuario='estudiante'
                        )
                    
                    # 4. Asignar Curso
                    if nombre_curso:
                        try:
                            curso = Curso.objects.filter(nombre__iexact=str(nombre_curso).strip(), activo=True).first()
                            
                            if curso:
                                user = User.objects.get(username=rut)
                                InscripcionCurso.objects.get_or_create(
                                    estudiante=user,
                                    curso=curso,
                                    año=curso.año, # Año actual del curso
                                    defaults={'estado': 'activo'}
                                )
                            else:
                                errores_log.append(f"Fila {row_idx}: Curso '{nombre_curso}' no encontrado.")
                        except Exception as e_curso:
                            errores_log.append(f"Fila {row_idx}: Error asignando curso ({str(e_curso)})")

                    creados += 1

                except Exception as e:
                    errores += 1
                    errores_log.append(f"Fila {row_idx} (RUT {rut}): {str(e)}")

            messages.success(request, f"Proceso finalizado. Registros procesados: {creados}. Errores: {errores}")
            
            if errores_log:
                for err in errores_log[:5]:
                    messages.warning(request, err)

            return redirect('administrativo:carga_masiva_estudiantes')
    else:
        form = CargaMasivaForm()

    return render(request, 'administrativo/carga_masiva.html', {
        'form': form,
        'page_title': 'Carga Masiva de Alumnos'
    })

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def carga_masiva_profesores(request):
    """Procesa la carga masiva de profesores desde Excel"""
    from .forms import CargaMasivaForm
    import openpyxl
    from django.contrib.auth.models import User, Group
    from usuarios.models import PerfilUsuario

    if request.method == 'POST':
        form = CargaMasivaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active

                creados = 0
                errores = 0
                errores_log = []

                # Iterar filas (saltando encabezado)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    # Protección contra filas vacías
                    if not row or all(cell is None for cell in row):
                        continue
                        
                    # Desempaquetado tolerante
                    try:
                        rut = row[0]
                        nombres = row[1]
                        apellidos = row[2]
                        email = row[3] if len(row) > 3 else None
                    except IndexError:
                        continue # Fila inválida
                    
                    if not rut or not nombres or not apellidos:
                        continue
                    
                    rut = str(rut).strip().upper().replace('.', '')
                    
                    try:
                        # 1. Verificar si usuario ya existe
                        if User.objects.filter(username=rut).exists():
                            pass
                        else:
                            # 2. Crear Usuario
                            raw_password = secrets.token_urlsafe(8)
                            user = User.objects.create_user(
                                username=rut,
                                email=email or '',
                                password=raw_password,
                                first_name=nombres,
                                last_name=apellidos,
                                is_staff=False # No staff por defecto, o True si se desea que accedan al admin
                            )
                            
                            # 3. Asignar Grupo Profesores
                            try:
                                group = Group.objects.get(name='Profesores')
                                user.groups.add(group)
                            except Group.DoesNotExist:
                                pass # O manejar creando el grupo

                            # 4. Crear Perfil
                            PerfilUsuario.objects.create(
                                user=user,
                                rut=rut,
                                tipo_usuario='profesor'
                            )
                            
                            creados += 1

                    except Exception as e:
                        errores += 1
                        errores_log.append(f"Fila {row_idx} (RUT {rut}): {str(e)}")

                messages.success(request, f"Carga de profesores finalizada. Registros: {creados}. Errores: {errores}")
                
                if errores_log:
                    for err in errores_log[:5]:
                        messages.warning(request, err)
                        
                return redirect('administrativo:carga_masiva_profesores')

            except Exception as e:
                messages.error(request, f"Error procesando archivo: {str(e)}")
    else:
        form = CargaMasivaForm()

    return render(request, 'administrativo/carga_masiva_profesores.html', {
        'form': form,
        'page_title': 'Carga Masiva de Profesores'
    })


@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def historial_actividad(request):
    """Vista completa del Log de Auditoría"""
    from .models import RegistroActividad
    from django.core.paginator import Paginator
    from usuarios.models import PerfilUsuario
    
    # Filtros
    tipo = request.GET.get('tipo', '')
    usuario_id = request.GET.get('usuario', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    
    qs = RegistroActividad.objects.select_related('usuario__perfil').all()
    
    if tipo:
        qs = qs.filter(tipo_accion=tipo)
    
    if usuario_id:
        qs = qs.filter(usuario_id=usuario_id)
        
    if fecha_inicio:
        qs = qs.filter(fecha__date__gte=fecha_inicio)
        
    paginator = Paginator(qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Lista de profesores con selection logic
    profesores_list = []
    qs_profesores = PerfilUsuario.objects.filter(tipo_usuario='profesor').select_related('user').order_by('user__last_name')
    target_usuario_id = int(usuario_id) if usuario_id else None
    
    for p in qs_profesores:
        p.is_selected = (p.user.id == target_usuario_id)
        profesores_list.append(p)
        
    # Tipos de acción con selection logic
    tipos_accion_list = []
    for valor, texto in RegistroActividad.TIPO_ACCION_CHOICES:
        tipos_accion_list.append({
            'valor': valor,
            'texto': texto,
            'is_selected': (valor == tipo)
        })
    
    return render(request, 'administrativo/historial_actividad.html', {
        'page_obj': page_obj,
        'tipos_accion_list': tipos_accion_list,
        'profesores': profesores_list,
        'filtros': {
            'tipo': tipo,
            'usuario': target_usuario_id,
            'fecha_inicio': fecha_inicio
        }
    })

@login_required
@user_passes_test(es_administrativo_check, login_url='usuarios:login')
def monitor_recursos(request):
    """
    Vista para que administradores monitoreen el contenido subido por profesores.
    Muestra: Recursos Académicos.
    Filtros: Profesor, Curso.
    """
    from academico.models import RecursoAcademico, Curso
    from usuarios.models import PerfilUsuario
    from django.core.paginator import Paginator
    
    # Filtros
    profesor_id = request.GET.get('profesor', '')
    curso_id = request.GET.get('curso', '')
    
    # QueryBase
    recursos = RecursoAcademico.objects.select_related('profesor', 'curso', 'asignatura').all().order_by('-creado')
    
    if profesor_id:
        recursos = recursos.filter(profesor_id=profesor_id)
    
    if curso_id:
        recursos = recursos.filter(curso_id=curso_id)
        
    # Paginación
    paginator = Paginator(recursos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Contexto para filtros
    # Contexto para filtros con lógica de selección pre-calculada (bypassing template quirks)
    profesores_list = []
    profesores_qs = PerfilUsuario.objects.filter(tipo_usuario='profesor').select_related('user').order_by('user__last_name')
    target_profesor_id = int(profesor_id) if profesor_id else None
    
    for p in profesores_qs:
        p.is_selected = (p.user.id == target_profesor_id)
        profesores_list.append(p)

    cursos_list = []
    # Admin debe ver todos los cursos, incluso inactivos, para auditar recursos pasados
    cursos_qs = Curso.objects.all().order_by('nivel', 'letra')
    target_curso_id = int(curso_id) if curso_id else None
    
    for c in cursos_qs:
        c.is_selected = (c.id == target_curso_id)
        cursos_list.append(c)
    
    return render(request, 'administrativo/monitor_recursos.html', {
        'page_obj': page_obj,
        'profesores': profesores_list,
        'cursos': cursos_list,
        'filtros': {
            'profesor': target_profesor_id,
            'curso': target_curso_id
        },
        'page_title': 'Monitor de Recursos Académicos'
    })
