from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import User

from django.core.exceptions import ValidationError
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
import csv
import io
import openpyxl
from .models import PerfilUsuario, Pupilo
from .forms import QuickStudentCreationForm, ImportacionMasivaForm
from academico.models import InscripcionCurso, Curso
from core.models import ConfiguracionAcademica
from django.http import HttpResponseRedirect
from django.urls import reverse


# ====================================
# Admin para gestión de Pupilos
# ====================================
@admin.register(Pupilo)
class PupiloAdmin(admin.ModelAdmin):
    """Admin para gestionar la relación Apoderado-Estudiante"""
    list_display = ('estudiante_nombre', 'apoderado_nombre', 'vinculo', 'es_apoderado_principal', 'created_at')
    list_filter = ('vinculo', 'es_apoderado_principal')
    search_fields = (
        'estudiante__user__first_name', 
        'estudiante__user__last_name',
        'estudiante__rut',
        'apoderado__user__first_name',
        'apoderado__user__last_name',
        'apoderado__rut'
    )
    raw_id_fields = ('apoderado', 'estudiante')
    list_per_page = 25
    
    fieldsets = (
        ('Relación', {
            'fields': ('apoderado', 'estudiante', 'vinculo')
        }),
        ('Estado', {
            'fields': ('es_apoderado_principal',)
        }),
    )
    
    def estudiante_nombre(self, obj):
        return obj.estudiante.nombre_completo
    estudiante_nombre.short_description = 'Estudiante'
    
    def apoderado_nombre(self, obj):
        return obj.apoderado.nombre_completo
    apoderado_nombre.short_description = 'Apoderado'


class PerfilUsuarioInline(admin.StackedInline):
    model = PerfilUsuario
    can_delete = False
    verbose_name_plural = 'Perfil'
    fields = ('rut', 'tipo_usuario', 'telefono', 'direccion', 'fecha_nacimiento', 'activo')

@admin.register(PerfilUsuario)
class PerfilUsuarioAdmin(admin.ModelAdmin):
    form = QuickStudentCreationForm
    list_display = ('rut', 'nombre_completo', 'tipo_usuario', 'email_usuario', 'activo', 'created_at')
    list_filter = ('tipo_usuario', 'activo', 'created_at')
    search_fields = ('rut', 'user__first_name', 'user__last_name', 'user__email', 'user__username')
    readonly_fields = ('created_at', 'updated_at')
    list_per_page = 25
    list_editable = ('activo',)
    
    change_list_template = "admin/usuarios/perfilusuario/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('importar-usuarios/', self.admin_site.admin_view(self.importacion_masiva), name='usuarios_perfilusuario_import_csv'),
        ]
        return my_urls + urls

    def importacion_masiva(self, request):
        if request.method == "POST":
            form = ImportacionMasivaForm(request.POST, request.FILES)
            if form.is_valid():
                archivo = request.FILES["archivo"]
                curso = form.cleaned_data['curso']
                
                estudiantes_data = []
                errors = []
                
                # Detectar formato
                if archivo.name.endswith('.xlsx'):
                    try:
                        wb = openpyxl.load_workbook(archivo)
                        ws = wb.active
                        headers = [cell.value for cell in ws[1]]
                        
                        # Mapeo de columnas
                        col_map = {}
                        for idx, header in enumerate(headers):
                            if header:
                                col_map[header.lower()] = idx
                                
                        # Verificar columnas requeridas
                        required = ['rut', 'nombres', 'apellidos']
                        missing = [req for req in required if req not in col_map]
                        
                        if missing:
                            messages.error(request, f"Faltan columnas requeridas en el Excel: {', '.join(missing)}")
                            return redirect("admin:usuarios_perfilusuario_changelist")
                            
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not row[col_map['rut']]: continue
                            
                            estudiantes_data.append({
                                'RUT': str(row[col_map['rut']]),
                                'Nombres': str(row[col_map['nombres']]),
                                'Apellidos': str(row[col_map['apellidos']]),
                                'Email': str(row[col_map.get('email')]) if 'email' in col_map and row[col_map['email']] else ''
                            })
                            
                    except Exception as e:
                        messages.error(request, f"Error leyendo Excel: {str(e)}")
                        return redirect("admin:usuarios_perfilusuario_changelist")
                        
                else: # Asumir CSV
                    try:
                        # Intentar decodificar
                        decoded_file = None
                        encodings = ['utf-8-sig', 'latin-1', 'cp1252']
                        file_data = archivo.read()
                        
                        for encoding in encodings:
                            try:
                                decoded_file = file_data.decode(encoding)
                                break
                            except UnicodeDecodeError:
                                continue
                                
                        if not decoded_file:
                            messages.error(request, "No se pudo decodificar el archivo CSV.")
                            return redirect("admin:usuarios_perfilusuario_changelist")

                        io_string = io.StringIO(decoded_file)
                        reader = csv.DictReader(io_string, delimiter=';')
                        
                        if 'RUT' not in reader.fieldnames:
                            io_string.seek(0)
                            reader = csv.DictReader(io_string, delimiter=',')
                            
                        estudiantes_data = list(reader)
                        
                    except Exception as e:
                        messages.error(request, f"Error leyendo CSV: {str(e)}")
                        return redirect("admin:usuarios_perfilusuario_changelist")

                # Procesar datos
                created_count = 0
                
                for row in estudiantes_data:
                    try:
                        # Normalizar keys (CSV vs Excel manual map)
                        # Si viene de Excel ya lo normalizamos, si es CSV keys son case-sensitive
                        rut = row.get('RUT') or row.get('rut')
                        nombres = row.get('Nombres') or row.get('nombres')
                        apellidos = row.get('Apellidos') or row.get('apellidos')
                        email = row.get('Email') or row.get('email') or ''
                        
                        if not rut or not nombres or not apellidos:
                            continue
                            
                        rut = rut.strip()
                        
                        # Check if user exists
                        if User.objects.filter(username=rut).exists():
                            # Si existe, verificar si está en el curso
                            user = User.objects.get(username=rut)
                            anio_actual = ConfiguracionAcademica.get_actual().año_actual
                            if not InscripcionCurso.objects.filter(estudiante=user, curso=curso, año=anio_actual).exists():
                                InscripcionCurso.objects.create(estudiante=user, curso=curso, año=anio_actual, estado='activo')
                                created_count += 1
                            continue
                            
                        # Create User
                        rut_limpio = rut.replace('.', '')
                        password = rut_limpio[:6] 
                        
                        if not email:
                            email = f"{rut}@liceo.cl"
                            
                        user = User.objects.create_user(
                            username=rut,
                            password=password,
                            first_name=nombres,
                            last_name=apellidos,
                            email=email,
                            is_staff=False
                        )
                        
                        PerfilUsuario.objects.create(
                            user=user,
                            rut=rut,
                            tipo_usuario='estudiante',
                            activo=True
                        )
                        
                        anio_actual = ConfiguracionAcademica.get_actual().año_actual
                        InscripcionCurso.objects.create(
                            estudiante=user,
                            curso=curso,
                            año=anio_actual,
                            estado='activo'
                        )
                        
                        created_count += 1
                        
                    except Exception as e:
                        errors.append(f"Error con {rut}: {str(e)}")
                
                if created_count > 0:
                    messages.success(request, f"Se procesaron {created_count} estudiantes exitosamente.")
                
                if errors:
                    messages.warning(request, f"Hubo errores: {'; '.join(errors[:5])}")
                    
                return redirect("admin:usuarios_perfilusuario_changelist")
        else:
            form = ImportacionMasivaForm()
            
        context = {
            'form': form,
            'opts': self.model._meta,
            'title': 'Importación Masiva de Alumnos',
            'site_title': self.admin_site.site_title,
            'site_header': self.admin_site.site_header,
            'has_permission': True,
        }
        return render(request, "admin/usuarios/perfilusuario/import_csv.html", context)

    fieldsets = (
        ('Información Personal', {
            'fields': ('rut', 'first_name', 'last_name', 'tipo_usuario', 'fecha_nacimiento')
        }),
        ('Credenciales de Acceso', {
            'fields': ('email', 'password1', 'password2'),
            'description': 'Configuración del usuario para iniciar sesión en la página web.'
        }),
        ('Contacto', {
            'fields': ('telefono', 'direccion')
        }),
        ('Estado', {
            'fields': ('activo',)
        }),
    )

    def get_fieldsets(self, request, obj=None):
        if not obj:
            return self.fieldsets
        
        return (
            ('Información Personal', {
                'fields': ('rut', 'tipo_usuario', 'fecha_nacimiento')
            }),
            ('Contacto', {
                'fields': ('telefono', 'direccion')
            }),
            ('Estado', {
                'fields': ('activo',)
            }),
        )

    def get_form(self, request, obj=None, **kwargs):
        Form = super().get_form(request, obj, **kwargs)
        if obj:
            class PerfilChangeForm(Form):
                def __init__(self, *args, **kwargs):
                    super().__init__(*args, **kwargs)
                    # Hacer campos de contraseña opcionales en edición
                    if 'password1' in self.fields:
                         self.fields['password1'].required = False
                    if 'password2' in self.fields:
                         self.fields['password2'].required = False
                    # Eliminar nombre y email que ya están en User
                    fields_to_remove = ['first_name', 'last_name', 'email']
                    for field in fields_to_remove:
                        if field in self.fields:
                            del self.fields[field]
            return PerfilChangeForm
        return Form

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        
        # Procesar cambio de contraseña si se proporcionó
        if change:
            p1 = form.cleaned_data.get('password1')
            p2 = form.cleaned_data.get('password2')
            
            if p1 and p2 and p1 == p2:
                try:
                    user = obj.user
                    user.set_password(p1)
                    user.save()
                    messages.success(request, f"Contraseña actualizada para {user.username}")
                except Exception as e:
                    messages.error(request, f"Error actualizando contraseña: {e}")

    def nombre_completo(self, obj):
        return obj.nombre_completo
    nombre_completo.short_description = 'Nombre Completo'
    
    def email_usuario(self, obj):
        return obj.user.email
    email_usuario.short_description = 'Email'

    def delete_model(self, request, obj):
        # Eliminar el usuario asociado (el perfil se elimina por CASCADE)
        obj.user.delete()

    def delete_queryset(self, request, queryset):
        # Eliminar los usuarios asociados para cada perfil seleccionado
        for perfil in queryset:
            perfil.user.delete()

    @admin.action(description='Inscribir seleccionados en un Curso')
    def inscribir_en_curso(self, request, queryset):
        # Si ya se envió el formulario con el curso seleccionado
        if 'apply' in request.POST:
            curso_id = request.POST.get('curso')
            if not curso_id:
                self.message_user(request, "Debe seleccionar un curso.", level=messages.ERROR)
                return HttpResponseRedirect(request.get_full_path())
            
            curso = Curso.objects.get(pk=curso_id)
            anio_actual = ConfiguracionAcademica.get_actual().año_actual
            count = 0
            for perfil in queryset:
                # Verificar si ya está inscrito
                if not InscripcionCurso.objects.filter(estudiante=perfil.user, curso=curso, año=anio_actual).exists():
                    InscripcionCurso.objects.create(
                        estudiante=perfil.user,
                        curso=curso,
                        año=anio_actual,
                        estado='activo'
                    )
                    count += 1
            
            self.message_user(request, f"Se inscribieron {count} estudiantes en el curso {curso}.")
            return HttpResponseRedirect(reverse('admin:usuarios_perfilusuario_changelist'))

        # Si es la primera vez (GET), mostrar formulario intermedio
        cursos = Curso.objects.filter(activo=True)
        context = {
            'cursos': cursos,
            'profiles': queryset,
            'opts': self.model._meta,
            'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
        }
        return render(request, 'admin/usuarios/perfilusuario/inscribir_curso_intermediate.html', context)

    actions = ['inscribir_en_curso']

# Extender el admin de User para incluir el perfil
class UserAdmin(BaseUserAdmin):
    inlines = (PerfilUsuarioInline,)
    list_display = ('username', 'email', 'first_name', 'last_name', 'tipo_usuario', 'is_staff', 'is_active')
    list_filter = ('is_staff', 'is_superuser', 'is_active', 'perfil__tipo_usuario')
    search_fields = ('username', 'first_name', 'last_name', 'email')
    
    def tipo_usuario(self, obj):
        try:
            return obj.perfil.get_tipo_usuario_display()
        except Exception:
            return 'Sin perfil'
    tipo_usuario.short_description = 'Tipo de Usuario'

    @admin.action(description='Inscribir seleccionados en un Curso')
    def inscribir_usuarios_en_curso(self, request, queryset):
        if 'apply' in request.POST:
            curso_id = request.POST.get('curso')
            if not curso_id:
                self.message_user(request, "Debe seleccionar un curso.", level=messages.ERROR)
                return HttpResponseRedirect(request.get_full_path())
            
            curso = Curso.objects.get(pk=curso_id)
            anio_actual = ConfiguracionAcademica.get_actual().año_actual
            count = 0
            for user in queryset:
                # Verificar si es estudiante
                if not hasattr(user, 'perfil') or user.perfil.tipo_usuario != 'estudiante':
                    continue
                    
                if not InscripcionCurso.objects.filter(estudiante=user, curso=curso, año=anio_actual).exists():
                    InscripcionCurso.objects.create(
                        estudiante=user,
                        curso=curso,
                        año=anio_actual,
                        estado='activo'
                    )
                    count += 1
            
            self.message_user(request, f"Se inscribieron {count} estudiantes en el curso {curso}.")
            return HttpResponseRedirect(reverse('admin:auth_user_changelist'))

        cursos = Curso.objects.filter(activo=True)
        context = {
            'cursos': cursos,
            'users': queryset,
            'opts': self.model._meta,
            'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
        }
        return render(request, 'admin/usuarios/user_inscribir_curso.html', context)

    actions = ['inscribir_usuarios_en_curso']

# Re-registrar User con el nuevo admin
admin.site.unregister(User)
admin.site.register(User, UserAdmin)

# Personalizar el admin site
admin.site.site_header = "Administración - Liceo Juan Bautista de Hualqui"
admin.site.site_title = "Administración del Liceo"
admin.site.index_title = "Panel de Administración"
